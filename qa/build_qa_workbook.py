from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


OUT_FILE = Path(__file__).with_name("HappyMeals_QA_TestCase_Defect_v2.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Тест-план"
for name in ["Чек-лист + Дефекты", "Тест-кейс 1", "Дефект", "Отчет"]:
    wb.create_sheet(name)

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="366092")
gray_fill = PatternFill("solid", fgColor="BFBFBF")
green_fill = PatternFill("solid", fgColor="92D050")
red_fill = PatternFill("solid", fgColor="FF0000")


def apply_header(cell):
    cell.fill = header_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(wrap_text=True, vertical="bottom")
    cell.border = border


def border_range(sheet, cell_range):
    for row in sheet[cell_range]:
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="center")


def set_common_alignment(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                old = copy(cell.alignment)
                cell.alignment = Alignment(
                    horizontal=old.horizontal,
                    vertical=old.vertical or "center",
                    wrap_text=True,
                )


today = date.today().strftime("%d.%m.%Y")

# Тест-план
ws = wb["Тест-план"]
for col, width in {
    "A": 9,
    "B": 44,
    "C": 5,
    "D": 33,
    "E": 9,
    "F": 13,
    "G": 17,
    "H": 10,
    "I": 14,
    "J": 13,
    "K": 20,
}.items():
    ws.column_dimensions[col].width = width

ws["D2"] = "Тест-план по системному тестированию Шеф-помощник"
ws["D2"].font = Font(bold=True, size=14)
ws["D2"].alignment = Alignment(horizontal="center")
ws["B3"] = "Цели доработки"
ws["B3"].font = Font(bold=True)
ws["B4"] = (
    "Обеспечение корректной работы клиент-серверного приложения помощника шефа: "
    "авторизация, фильтрация рецептов, история поиска, избранное и статистика пользователя"
)
ws["B6"] = (
    "Важен критерий: задокументированы все дефекты, исправлены все дефекты "
    "с приоритетом выше Critical"
)
ws["B8"] = "1 - самый высокий приоритет"

headers = [
    "JiraTask",
    "Область функционала",
    "Прио",
    "Стратегия тестирования",
    "h",
    "Риски",
    "Статус",
    "Аналитик",
    "Разработчик",
    "Тестировщик",
    "FSD",
]
for idx, header in enumerate(headers, 1):
    apply_header(ws.cell(9, idx, header))

plan_rows = [
    [
        "",
        "Регистрация и авторизация",
        1,
        "Проверить валидацию логина, пароля, email, запрет дублей логина и вход только по существующим пользователям",
        1,
        "",
        "протестирован, есть ошибки",
        "",
        "",
        "",
        "",
    ],
    [
        "",
        "Фильтры рецептов",
        1,
        "Проверить корректность фильтрации по ингредиентам, кухне, типу блюда, времени и сложности",
        1,
        "",
        "протестирован, есть ошибки",
        "",
        "",
        "",
        "",
    ],
    [
        "",
        "История поиска",
        1,
        "Проверить читаемость истории, отсутствие дублей и привязку истории к аккаунту",
        0.5,
        "",
        "протестирован, есть ошибки",
        "",
        "",
        "",
        "",
    ],
    [
        "",
        "Избранное и статистика",
        1,
        "Проверить сохранение избранного и статистики по user_id, включая время в приложении",
        0.5,
        "",
        "протестирован, есть ошибки",
        "",
        "",
        "",
        "",
    ],
]
for row_idx, row in enumerate(plan_rows, 10):
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row_idx, col_idx, value)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if col_idx in (3, 5):
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if col_idx == 3:
            cell.font = Font(color="FF0000")
        if col_idx == 5:
            cell.font = Font(bold=True)
ws["A14"] = "Итого"
ws["A14"].font = Font(bold=True)
ws["E14"] = "=SUM(E10:E13)"
ws["E14"].font = Font(bold=True)
border_range(ws, "A9:K14")
ws.freeze_panes = "A9"

# Чек-лист + Дефекты
ws = wb["Чек-лист + Дефекты"]
for col, width in {
    "A": 13,
    "B": 28,
    "C": 42,
    "D": 12,
    "E": 12,
    "F": 14,
    "J": 58,
    "K": 14,
}.items():
    ws.column_dimensions[col].width = width
ws["A1"] = "Тест-кейсы"
ws["F1"] = "Результат"
ws["F2"] = today
for cell_ref in ("A1", "F1", "F2"):
    ws[cell_ref].fill = gray_fill
    ws[cell_ref].font = Font(bold=True)
    ws[cell_ref].border = border
    ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")

check_rows = [
    ("Registration", "Register valid user", "Создать пользователя с валидным логином, паролем и email", ""),
    ("Registration", "Reject numeric login", "Попробовать зарегистрировать логин, состоящий только из цифр", ""),
    ("Registration", "Reject short password", "Попробовать зарегистрировать пароль короче 8 символов", ""),
    ("Registration", "Reject invalid email", "Попробовать зарегистрировать email без формата example@example.example", ""),
    ("Registration", "Reject duplicate login", "Повторить регистрацию с уже существующим логином", ""),
    ("Registration", "Redirect to login after register", "После успешной регистрации проверить переход на форму входа", ""),
    ("Authorization", "Login valid user", "Войти под существующим пользователем с корректным паролем", ""),
    ("Authorization", "Reject wrong password", "Ввести существующий логин и неверный пароль", ""),
    ("Authorization", "Reject unknown login", "Ввести несуществующий логин и произвольный пароль", ""),
    ("Authorization", "Logout clears session", "Выйти из аккаунта и проверить очистку данных текущей сессии", ""),
    ("Server connection", "Connect to remote server", "Запустить клиент с IP и портом удаленного сервера", ""),
    ("Server connection", "Handle unavailable server", "Запустить клиент при недоступном сервере и проверить сообщение об ошибке", ""),
    ("Filters", "Search without filters", "Выполнить поиск без ограничений", ""),
    ("Filters", "Search by cuisine", "Подобрать рецепты по выбранной кухне", ""),
    ("Filters", "Search by dish type", "Подобрать рецепты по типу блюда", ""),
    ("Filters", "Search by max time", "Подобрать рецепты по максимальному времени приготовления", ""),
    ("Filters", "Search by complexity", "Подобрать рецепты по сложности", ""),
    ("Filters", "Search by excluded ingredient", "Исключить ингредиент и проверить, что рецепты с ним не выводятся", ""),
    ("Filters", "Combined filters", "Проверить совместную работу кухни, типа, времени, сложности и исключений", ""),
    ("Results", "Render result list", "Проверить, что найденные блюда выводятся отдельными элементами списка", ""),
    ("Results", "No results state", "Задать фильтры без совпадений и проверить сообщение 'Рецепты не найдены'", ""),
    ("Results", "Clear previous selected dish", "После нового поиска без результатов проверить очистку старого выбранного блюда", ""),
    ("Recipe details", "Open recipe details", "Открыть найденное блюдо и проверить название, кухню, время и описание", ""),
    ("Recipe details", "Back to filters", "Вернуться из описания рецепта к фильтрам", ""),
    ("Favorites", "Add favorite", "Добавить рецепт в избранное", ""),
    ("Favorites", "Prevent favorite duplicate", "Повторно добавить тот же рецепт в избранное", ""),
    ("Favorites", "Remove favorite", "Удалить рецепт из избранного", ""),
    ("Favorites", "Favorites account binding", "Войти под другим пользователем и проверить отдельное избранное", ""),
    ("History", "Save search history", "Выполнить поиск и проверить появление записи в истории", ""),
    ("History", "Readable history", "Открыть историю и проверить человекочитаемый формат записи", "1"),
    ("History", "No duplicate history", "Повторить одинаковый поиск и проверить отсутствие подряд идущих дублей", "1"),
    ("History", "History account binding", "Войти под другим пользователем и проверить отдельную историю", ""),
    ("Statistics", "Search count", "Проверить увеличение счетчика поисковых запросов", ""),
    ("Statistics", "Favorites count", "Проверить изменение счетчика избранных рецептов", ""),
    ("Statistics", "Total app time", "Проверить счетчик общего времени в приложении", ""),
    ("Statistics", "Statistics account binding", "Войти под другим пользователем и проверить отдельную статистику", ""),
]
for row_idx, row in enumerate(check_rows, 3):
    area, name, description, defect_no = row
    ws.cell(row_idx, 1, area)
    ws.cell(row_idx, 2, name)
    ws.cell(row_idx, 3, description)
    ws.cell(row_idx, 6, defect_no)
    for col_idx in range(1, 7):
        cell = ws.cell(row_idx, col_idx)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.cell(row_idx, 6).fill = red_fill if defect_no else green_fill
    if defect_no:
        ws.cell(row_idx, 6).font = Font(bold=True)

ws.merge_cells("J1:K1")
ws["J1"] = "Дефект"
ws["J1"].fill = gray_fill
ws["J1"].font = Font(bold=True)
ws["J1"].alignment = Alignment(horizontal="center")
ws["J2"] = "№ Наименование"
ws["K2"] = "Важность"
for cell_ref in ("J1", "K1", "J2", "K2"):
    ws[cell_ref].border = border
    ws[cell_ref].font = Font(bold=True)
    ws[cell_ref].fill = gray_fill
    ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
ws["J3"] = "1 История поиска отображается технической строкой и дублируется после повторного поиска"
ws["K3"] = "High"
for row_idx in range(3, 42):
    for col_idx in range(10, 12):
        ws.cell(row_idx, col_idx).border = border
        ws.cell(row_idx, col_idx).alignment = Alignment(wrap_text=True, vertical="center")

# Тест-кейс 1
ws = wb["Тест-кейс 1"]
for col, width in {
    "A": 6,
    "B": 29,
    "C": 36,
    "D": 8,
    "E": 8,
    "F": 8,
    "G": 31,
    "H": 13,
}.items():
    ws.column_dimensions[col].width = width
for cell_ref in ("B1", "B2", "B3", "B4", "D1", "D2", "D3", "D4"):
    ws[cell_ref].fill = gray_fill
    ws[cell_ref].font = Font(bold=True)
    ws[cell_ref].border = border
    ws[cell_ref].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
ws["B1"] = "Наименование:"
ws["C1"] = "История поиска. Читаемость и отсутствие дублей"
ws["B2"] = "Описание:"
ws["C2"] = "Тест-кейс для проверки, что история поиска сохраняется в понятном виде и не дублируется после повторного одинакового запроса"
ws["D1"] = "№:"
ws["G1"] = "TC-HM-001"
ws["D2"] = "Статус:"
ws["G2"] = "Failed"
ws["D3"] = "Дефекты №:"
ws["G3"] = "1"
ws["B4"] = "Тестировщик:"
ws["D4"] = "Дата:"
ws["G4"] = today
for cell_ref in ("C1", "C2", "C3", "C4", "G1", "G2", "G3", "G4", "H1", "H2", "H3", "H4"):
    ws[cell_ref].border = border
    ws[cell_ref].alignment = Alignment(wrap_text=True, vertical="center")
ws["B6"] = "Начальные условия:"
ws["B6"].font = Font(bold=True)
ws["B6"].alignment = Alignment(horizontal="right")
ws["C6"] = "Сервер запущен, пользователь зарегистрирован и авторизован, база рецептов доступна"
ws["B7"] = "<список параметров>"
ws["C7"] = "<соответствующие значения>"
ws["D7"] = "<отметка о проверке при выполнении кейса>"
border_range(ws, "B6:G7")
for col_idx, value in enumerate(["5", "шагов", "Число шагов по статусам:", "4", "0", "1", "% Complete:", "80%"], 1):
    cell = ws.cell(10, col_idx, value)
    cell.fill = gray_fill
    cell.border = border
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for col_idx, value in enumerate(["№ п/п", "Действие", "Ожидаемый результат", "Pass", "Fail", "N/A", "Фактический результат", "№ дефекта"], 1):
    cell = ws.cell(11, col_idx, value)
    cell.fill = gray_fill
    cell.font = Font(bold=True)
    cell.border = border
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
steps = [
    (1, "Запустить клиент HappyMealsClient и войти под пользователем user123", "Открывается главное меню приложения", "x", "", "", "Открыто главное меню", ""),
    (2, "Открыть форму фильтров и выбрать: кухня RUSSIAN, время до 60 минут", "Параметры фильтра выбраны без ошибок", "x", "", "", "Параметры выбраны", ""),
    (3, "Нажать поиск, затем открыть найденный рецепт и вернуться к фильтрам", "Список рецептов отображается отдельными блюдами", "x", "", "", "Рецепты отображены", ""),
    (4, "Повторить такой же поиск и открыть История поисков", "В истории одна читаемая запись вида: кухня: RUSSIAN; время до 60 мин; без технических ключей и дублей", "", "x", "", "Отображаются строки вида ingredients=any; cuisines=RUSSIAN; maxTime=60, есть повторяющиеся записи", "1"),
    (5, "Выйти из аккаунта, войти под другим пользователем и открыть историю", "История второго пользователя не содержит запросы первого пользователя", "x", "", "", "История проверяется отдельно для аккаунта", ""),
]
for row_idx, row in enumerate(steps, 12):
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row_idx, col_idx, value)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for col_idx in (4, 5, 6):
        ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="center", vertical="center")
for row_idx, height in {12: 36, 13: 48, 14: 48, 15: 75, 16: 60}.items():
    ws.row_dimensions[row_idx].height = height

# Дефект
ws = wb["Дефект"]
for col, width in {"A": 20, "B": 25, "C": 20, "D": 22}.items():
    ws.column_dimensions[col].width = width
fields = [
    ("A1", "Название", "B1", "История поиска отображается технической строкой и дублируется"),
    ("A2", "№ дефекта", "B2", "1"),
    ("C2", "№ тест-кейса", "D2", "TC-HM-001"),
    ("A3", "Проект", "B3", "Шеф-помощник / HappyMeals"),
    ("C3", "Компонент", "D3", "История поиска"),
    ("A4", "Статус", "B4", "Opened"),
    ("C4", "Номер версии", "D4", "0.1"),
    ("A5", "Важность", "B5", "High"),
    ("C5", "Приоритет", "D5", "High"),
    ("A10", "Назначен на", "B10", ""),
    ("C10", "Автор", "D10", ""),
]
for label_cell, label, value_cell, value in fields:
    ws[label_cell] = label
    ws[value_cell] = value
    ws[label_cell].fill = gray_fill
    ws[label_cell].font = Font(bold=True)
    ws[label_cell].alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    ws[label_cell].border = border
    ws[value_cell].border = border
    ws[value_cell].alignment = Alignment(wrap_text=True, vertical="center")
for cell_ref, value in {
    "A6": "Blocker",
    "B6": "Critical",
    "A7": "Critical",
    "B7": "High",
    "A8": "Major",
    "B8": "Medium",
    "A9": "Minor",
    "B9": "Low",
    "C6": "Opened",
    "D6": "In progress",
    "C7": "Retest",
    "D7": "Fixed",
    "C8": "Closed",
}.items():
    ws[cell_ref] = value
for row_idx in range(6, 10):
    for col_idx in range(1, 5):
        ws.cell(row_idx, col_idx).border = border
ws["A11"] = "Описание"
ws["A11"].fill = gray_fill
ws["A11"].font = Font(bold=True)
ws["A11"].alignment = Alignment(horizontal="right", vertical="top")
ws["A11"].border = border
ws.merge_cells("B11:D13")
ws["B11"] = (
    "Шаги воспроизведения:\n"
    "1. Запустить сервер и клиент.\n"
    "2. Авторизоваться под пользователем user123.\n"
    "3. Открыть фильтры и выполнить поиск с параметрами: кухня RUSSIAN, время до 60 минут.\n"
    "4. Повторить тот же поиск.\n"
    "5. Открыть раздел История поисков.\n\n"
    "Ожидаемый результат:\n"
    "История содержит одну понятную запись: кухня RUSSIAN, время до 60 мин, без технических ключей, без дублей.\n\n"
    "Фактический результат:\n"
    "История показывает технические строки вида ingredients=any; cuisines=RUSSIAN; maxTime=60 и повторяет одинаковые запросы."
)
ws["B11"].alignment = Alignment(wrap_text=True, vertical="top")
for row in ws["B11:D13"]:
    for cell in row:
        cell.border = border
for row_idx in (11, 12, 13):
    ws.row_dimensions[row_idx].height = 115
ws["A14"] = "Вложения"
ws["A14"].fill = gray_fill
ws["A14"].font = Font(bold=True)
ws["A14"].alignment = Alignment(horizontal="right")
ws["A14"].border = border
ws.merge_cells("B14:D14")
ws["B14"] = "Скриншот истории поиска с техническими строками и дублями"
ws["B14"].border = border

# Отчет
ws = wb["Отчет"]
for col, width in {"A": 34, "B": 16, "C": 12}.items():
    ws.column_dimensions[col].width = width
for cell_ref, value in {
    "A1": "Отчет о системном тестировании Шеф-помощник",
    "A2": "Версия",
    "B2": "0.1",
    "A3": "Сроки проведения тестирования",
    "B3": today,
    "A4": "Участники процесса",
    "A7": "к-во запланированных тестов",
    "B7": 36,
    "A8": "к-во выполненных тестов",
    "B8": 36,
    "A9": "к-во успешно выполненных тестов",
    "B9": 34,
    "A10": "к-во неуспешно выполненных тестов",
    "B10": 2,
    "A12": "к-во зарегистрированных ошибок",
    "B12": 1,
    "A13": "приоритета Critical",
    "B13": 0,
    "A14": "приоритета High",
    "B14": 1,
    "A15": "приоритета Medium",
    "B15": 0,
    "A16": "приоритета Minor",
    "B16": 0,
    "A17": "приоритета Trivial",
    "B17": 0,
    "A20": "Заключение:",
    "A21": "Система не рекомендуется для установки в прод до исправления дефекта High в истории поиска.",
}.items():
    ws[cell_ref] = value
ws["C8"] = "=B8/B7"
ws["C9"] = "=B9/B8"
ws["C10"] = "=B10/B8"
for cell_ref in ("C8", "C9", "C10"):
    ws[cell_ref].number_format = "0.00%"
for row_idx in range(7, 18):
    for col_idx in range(1, 4):
        ws.cell(row_idx, col_idx).border = border
ws["A1"].font = Font(bold=True)
ws["A20"].font = Font(bold=True)

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = True
    set_common_alignment(sheet)

wb.save(OUT_FILE)
print(OUT_FILE.resolve())
